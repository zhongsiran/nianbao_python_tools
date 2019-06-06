import pymysql
from openpyxl import load_workbook
from openpyxl import Workbook
from string import Template
import os
import time

class UpdateDatabase:
    def __init__(self):
        self.conn = ""
        self.cur = ""

    def get_conn(self):
        conn = pymysql.connect(host='localhost', port=3306, user='homestead', passwd='secret', db='nianbao', charset='utf8')
        return conn

    def query_all(self, cur, sql, args):
        cur.execute(sql, args)
        return cur.fetchall()

    # 读取XLSX文件
    def loadworkbook(self):
        try:
            # 代码中指定读取的文件，每次修改
            wb = load_workbook('年报系统最新导出表.xlsx')
            # 第一个工作表
            self.ws = wb.worksheets[0]
        except FileNotFoundError:
            print('当前目录没有找到“年报系统最新导出表.xlsx”')
            exit(0)

    def read_workbook_and_update(self):
        # 从第三行开始
        rows = self.ws[3:self.ws.max_row]
        # rows = self.ws[3:5]
        # 每行进行一次
        for row in rows:
            # try...except用于去除空格

            # row[1]是第2列，必须有注册号，不能为空，发现某行为空刚报错，终止整个程序
            try:
                r = ''.join(row[1].value.split())
            except AttributeError as e:
                if row[1].value != "":
                    r = row[1].value
                else:
                    print(row[1].value)
                    print(e)
                    print('存在注册号为空的情况，请检查后重新运行。')
                    exit(0)
                print("regnum:" + str(r))
                
            for cell in row:
                if isinstance(cell.value, str):
                    if cell.value.find("已公示")!=-1:
                        self.update_db('已公示', r)
                    elif cell.value.find("未填报")!=-1:
                        self.update_db('未填报', r)
            print("====next=====")
                

    def update_db(self, nian_bao_status, regnum):
        self.conn = self.get_conn()
        self.cur = self.conn.cursor()
        sql = 'UPDATE 2019_nianbao_corp SET `nian_bao_status` = "%s" WHERE `RegNum` = "%s"' %(nian_bao_status, str(regnum))
        print("SQL TO COMMIT IS:" + sql)
        self.query_all(self.cur, sql, None)
        self.conn.commit()

    def close_cur(self):
        self.cur.close()

    def close_conn(self):
        self.conn.close()

if __name__ == '__main__':
    ud=UpdateDatabase()
    ud.loadworkbook()
    ud.read_workbook_and_update()
    ud.close_cur()
    ud.close_conn()

    # read_mysql_to_xlsx('年报系统最新导出表.xlsx')
