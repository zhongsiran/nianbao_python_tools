import pymysql
from openpyxl import load_workbook
from openpyxl import Workbook
from string import Template
from datetime import date
import sys

class ExportMysqlToXlsx:
    def __init__(self):

        self.host = 'localhost'
        self.host_2 = '10.2.94.28'  # 合益所
        # self.host_2 = '10.21.201.175' # 花城所
        self.port = 3306
        self.user = 'homestead'
        self.passwd = 'secret'
        self.db = 'nianbao'

        self.conn = self.get_conn()
        self.cur = self.conn.cursor()
        # self.workbook = Workbook()
        # self.worksheet = self.workbook.active

        self.worksheet_headrow = ['注册号','名称','地址','登记电话','法定代表人','联络员','联络员电话']
        self.worksheet_headrow.extend(['历史电话记录','本年度电话记录','跟进情况'])
        self.worksheet_headrow.extend(['年报状态','电话情况','联系员电话情况','企业类型','片区'])
        self.db_table_name = '2019_nianbao_corp'
        self.results = tuple()
        self.row = 1

    def create_new_workbook(self):
        workbook = Workbook()
        worksheet = workbook.active
        if self.worksheet_headrow:
            for i, j in enumerate(self.worksheet_headrow):
                worksheet.cell(row=1, column=i+1, value=j)
        return workbook

    def get_conn(self):
        print('连接中......')
        try:
            conn = pymysql.connect(host=self.host, port=self.port, user=self.user, passwd=self.passwd, db=self.db, charset='utf8')
            print('连接本机数据库成功')
        except pymysql.err.OperationalError as e:
            print('没有在本机发现数据库，尝试连接%s' % self.host_2)
            try:
                conn = pymysql.connect(host=self.host_2, port=self.port, user=self.user, passwd=self.passwd, db=self.db, charset='utf8')
                print('连接成功')
            except pymysql.err.OperationalError as e:
                print('无法连接到服务器，请检查设置，程序即将退出')
                print('错误信息：')
                print(e)
                sys.exit()
        
        return conn

    def query_all(self, cur, sql, args):
        print('查询中......')
        cur.execute(sql, args)
        return cur.fetchall()

    def get_db_data(self):
        sql = '''SELECT 
        Regnum, CorpName, Addr, Phone, RepPerson, ContactPerson, ContactPhone,
        PhoneCallHistoryRecord, PhoneCallRecord, Status,
        nian_bao_status, phone_status, cphone_status, type, division
        FROM %s
        WHERE nian_bao_status = '未填报'
        ORDER BY division, type
        ''' % self.db_table_name
        results = self.query_all(self.cur, sql, None)
        if results:
            self.results = results
        else:
            print('查询不到记录')
            self.results = ()

    def save_to_xlsx(self):
        # 初始化首个表格
        current_workbook = self.create_new_workbook()
        result_row = 2
        first_row_idx = 0
        file_count = 1
        # 迭代查询数据，得到序号，内容
        print('共找到' + str(len(self.results)) + '条记录')
        for idx, content in enumerate(self.results):
            print(idx)
            # 非最后一行
            if idx +1 != len(self.results):
                # 写入行为2且序号为0，即查询内容的首行，必定直接写入
                if result_row == 2 and idx == 0:
                    print('first file first row')
                    # 首记录写入首行
                    for col_idx, col_content in enumerate(content):
                        current_workbook.active.cell(row=result_row, column=col_idx+1, value=col_content)
                    result_row = result_row + 1
                # 写入行为2且上个文件最后判断行不为0，新建立的文件，需要补回上一判断记录
                elif result_row == 2 and first_row_idx > 0:
                    print('new file first row')
                    # 后续文件首行补上个文件最后判断的记录
                    for col_idx, col_content in enumerate(self.results[first_row_idx]):
                        current_workbook.active.cell(row=result_row, column=col_idx+1, value=col_content)
                    result_row = result_row + 1
                    # 判断是否单行文件情况，上个文件最后判断记录可能自成一个文件，需要判断。
                    if self.results[first_row_idx][14] == content[14]:
                        if self.results[first_row_idx][13] == content[13]:
                            for col_idx, col_content in enumerate(content):
                                current_workbook.active.cell(row=result_row, column=col_idx+1, value=col_content)
                            result_row = result_row + 1
                        # 如自成一个文件，保存为单独的文件
                        else:
                            self.save_workbook(current_workbook, file_count, self.results[idx-1][14], self.results[idx-1][13])
                            # 另建新表
                            current_workbook = self.create_new_workbook()
                            result_row = 2
                            # 保存当前判断行，在新建文件中补写入
                            first_row_idx = idx
                            file_count += 1
                    else:
                        self.save_workbook(current_workbook, file_count, self.results[idx-1][14], self.results[idx-1][13])
                        # 另建新表
                        current_workbook = self.create_new_workbook()
                        result_row = 2
                        # 保存当前判断行，在新建文件中补写入
                        first_row_idx = idx
                        file_count += 1
                    # 符合14片区、13类型均与上一行相等时，写入当前的文件中。
                elif result_row > 2 and self.results[idx][14] == self.results[idx-1][14]:
                    if self.results[idx][13] == self.results[idx-1][13]:
                        print('current file continue...')
                        for col_idx, col_content in enumerate(content):
                            current_workbook.active.cell(row=result_row, column=col_idx+1, value=col_content)
                        result_row = result_row + 1
                    # 否则保存当前文件，另建新文件。
                    else:
                        self.save_workbook(current_workbook, file_count, self.results[idx-1][14], self.results[idx-1][13])
                        current_workbook = self.create_new_workbook()
                        result_row = 2
                        first_row_idx = idx
                        file_count += 1
                # 保存当前文件，另建新文件。
                else:
                    self.save_workbook(current_workbook, file_count, self.results[idx-1][14], self.results[idx-1][13])
                    current_workbook = self.create_new_workbook()
                    result_row = 2
                    first_row_idx = idx
                    file_count += 1
            # 迭代完成后保存当前文件
            elif self.results[idx-1][14] == content[14] and self.results[idx-1][13] == content[13]:
                print('last row')
                # 在当前文件继续写入
                for col_idx, col_content in enumerate(content):
                    current_workbook.active.cell(row=result_row, column=col_idx+1, value=col_content)
                self.save_workbook(current_workbook, file_count, self.results[idx][14], self.results[idx][13])
            elif self.results[idx-1][14] != content[14] or self.results[idx-1][13] != content[13]:
                self.save_workbook(current_workbook, file_count, self.results[idx-1][14], self.results[idx-1][13])
                file_count += 1
                current_workbook = self.create_new_workbook()
                result_row = 2
                for col_idx, col_content in enumerate(content):
                    current_workbook.active.cell(row=result_row, column=col_idx+1, value=col_content)
                self.save_workbook(current_workbook, file_count, content[14], content[13])

    def save_workbook(self, workbook, file_count, division, corp_type):
        now = date.today()
        today = "%d-%d-%d" %(now.year, now.month, now.day)
        save_name = str(file_count) + '-' + today + '-' +  division + '-' + corp_type + '-未报名单.xlsx'
        try:
            workbook.save(save_name)
            print('SAVE TO ' + save_name)
        except PermissionError as e:
            print(e)
            print('出错了：无法保存文件“%s”，请检查是否已经打开同名文件。'% save_name)
        workbook.close()
        
    def close_cur(self):
        self.cur.close()

    def close_conn(self):
        self.conn.close()

if __name__ == "__main__":
    E = ExportMysqlToXlsx()
    E.get_db_data()
    if E.results:
        E.save_to_xlsx()
        E.close_cur()
        E.close_conn()
