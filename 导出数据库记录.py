import pymysql
from openpyxl import load_workbook
from openpyxl import Workbook
from string import Template
from datetime import date

class ExportMysqlToXlsx:
    def __init__(self):
        # self.conn = self.get_conn()
        # self.cur = self.conn.cursor()
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet_headrow = ['注册号','名称','地址','登记电话','法定代表人','联络员','联络员电话']
        self.worksheet_headrow.extend(['历史电话记录','本年度电话记录','跟进情况'])
        self.worksheet_headrow.extend(['年报状态','电话情况','联系员电话情况'])
        self.db_table_name = '2019_nianbao_corp'
        self.results = tuple()
        self.row = 1
        
    def get_conn(self):
        conn = pymysql.connect(host='10.2.94.28', port=3306, user='homestead', passwd='secret', db='nianbao', charset='utf8')
        return conn

    def query_all(self, cur, sql, args):
        cur.execute(sql, args)
        return cur.fetchall()

    def get_db_data(self):
        sql = '''SELECT 
        Regnum, CorpName, Addr, Phone, RepPerson, ContactPerson,
        PhoneCallHistoryRecord, PhoneCallRecord, Status,
        nian_bao_status, phone_status, cphone_status
        FROM %s
        WHERE PhoneCallRecord != '' 
        LIMIT 5''' % self.db_table_name
        results = self.query_all(self.cur, sql, None)
        if results:
            self.results = results
        else:
            print('查询不到记录')
            exit(0)

    def save_to_xlsx(self):
        if self.worksheet_headrow:
            for i, j in enumerate(self.worksheet_headrow):
                self.worksheet.cell(row=1, column=i+1, value=j)
        for result in self.results:
            print(result)
            for column in result:
                print(column)
        self.save_workbook()

    def save_workbook(self):
        now = date.today()
        today = "%d-%d-%d" %(now.year, now.month, now.day)
        self.workbook.save(today + '导出记录表.xlsx')

    def close_cur(self):
        self.cur.close()

    def close_conn(self):
        self.conn.close()

if __name__ == "__main__":
    E = ExportMysqlToXlsx()
    # E.get_db_data()
    E.save_to_xlsx()
    # E.close_cur()
    # E.close_conn()